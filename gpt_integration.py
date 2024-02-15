#!/usr/bin/env python
# coding: utf-8

# Execute the script to save the DataFrame to an Excel file
from openai import OpenAI
import openpyxl

# Set the GPT-4 API key
client = OpenAI(
    # defaults to os.environ.get("OPENAI_API_KEY")
    api_key="Your api_key",
)

# Read an Excel file
workbook = openpyxl.load_workbook("Path to your file")
sheet = workbook.active

# Apply the range of the row and column
start_row = "number"
end_row = "number"
question_column = "number"
answer_column = "number"

# Repeat the execution of the command 
for row in range(start_row, end_row + 1):
    question = sheet.cell(row=row, column=question_column).value

    if question:
        # Answer the question using the GPT-4 API
        response = client.chat.completions.create(
        model="gpt-4-1106-preview",
        messages=[{"role": "system", "content": question}]
    )

        answer = response.choices[0].message.content

        # Write an answer next to the cell
        sheet.cell(row=row, column=answer_column).value = answer

# Save the edited Excel file
workbook.save("Path to your file")